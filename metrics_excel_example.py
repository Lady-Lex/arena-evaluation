#!/usr/bin/env python3
"""
Generate an Excel file that contains all metrics data.
- Tasks 1-5: include metrics, metrics_subject, and metrics_zone data.
- Use different background and font colors to distinguish tasks.
"""

import os
import pandas as pd
import glob
from datetime import datetime
from pathlib import Path
import warnings
import numpy as np
import ast
try:
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl is not installed, skipping color styling")
warnings.filterwarnings('ignore')

def format_datetime():
    now_time = datetime.now()
    return now_time.strftime("%Y-%m-%d_%H-%M-%S")

def get_task_colors():
    """Define background colors for different tasks."""
    return {
        'task_1': 'FFE6E6',  # light red
        'task_2': 'E6F3FF',  # light blue
        'task_3': 'E6FFE6',  # light green
        'task_4': 'FFF0E6',  # light orange
        'task_5': 'F0E6FF'   # light purple
    }

def get_task_font_colors():
    """Define font colors for different tasks."""
    return {
        'task_1': '8B0000',  # dark red
        'task_2': '000080',  # dark blue
        'task_3': '006400',  # dark green
        'task_4': 'FF8C00',  # dark orange
        'task_5': '4B0082'   # dark purple
    }

def calculate_path_smoothness(odom_file):
    """
    Compute path smoothness metrics based on odometry data:
    1. curvature_smoothness: total heading angle change
    2. acceleration_smoothness: total change in acceleration
    3. path_length: total path length
    4. avg_velocity: average velocity
    """
    try:
        # Read odometry data
        df_odom = pd.read_csv(odom_file)
        
        if len(df_odom) < 3:
            return {
                'curvature_smoothness': 0.0,
                'acceleration_smoothness': 0.0,
                'path_length': 0.0,
                'avg_velocity': 0.0,
                'max_curvature': 0.0,
                'avg_curvature': 0.0
            }
        
        # Parse position data
        positions = []
        velocities = []
        times = []
        
        for _, row in df_odom.iterrows():
            try:
                data_dict = ast.literal_eval(row['data'])
                pos = data_dict['position']
                vel = data_dict['velocity']
                
                positions.append([pos[0], pos[1], pos[2]])  # x, y, theta
                velocities.append([vel[0], vel[1], vel[2]])  # vx, vy, vtheta
                times.append(row['time'])
            except:
                continue
        
        if len(positions) < 3:
            return {
                'curvature_smoothness': 0.0,
                'acceleration_smoothness': 0.0,
                'path_length': 0.0,
                'avg_velocity': 0.0,
                'max_curvature': 0.0,
                'avg_curvature': 0.0
            }
        
        positions = np.array(positions)
        velocities = np.array(velocities)
        times = np.array(times)
        
        # Compute path length
        path_length = 0.0
        for i in range(1, len(positions)):
            dx = positions[i][0] - positions[i-1][0]
            dy = positions[i][1] - positions[i-1][1]
            path_length += np.sqrt(dx*dx + dy*dy)
        
        # Compute average velocity
        total_time = (times[-1] - times[0]) / 1e9  # convert to seconds
        avg_velocity = path_length / total_time if total_time > 0 else 0.0
        
        # Compute curvature smoothness (heading angle changes)
        curvature_changes = []
        curvatures = []
        
        for i in range(1, len(positions)):
            # Compute heading angle change
            theta_prev = positions[i-1][2]
            theta_curr = positions[i][2]
            
            # Handle angle wrap-around
            angle_diff = theta_curr - theta_prev
            if angle_diff > np.pi:
                angle_diff -= 2 * np.pi
            elif angle_diff < -np.pi:
                angle_diff += 2 * np.pi
            
            curvature_changes.append(abs(angle_diff))
            
            # Compute instantaneous curvature (based on change in velocity direction)
            if i > 1:
                vx_prev = velocities[i-1][0]
                vy_prev = velocities[i-1][1]
                vx_curr = velocities[i][0]
                vy_curr = velocities[i][1]
                
                # Compute velocity direction angle
                angle_prev = np.arctan2(vy_prev, vx_prev)
                angle_curr = np.arctan2(vy_curr, vx_curr)
                
                # Compute curvature (rate of change of angle)
                angle_diff_vel = angle_curr - angle_prev
                if angle_diff_vel > np.pi:
                    angle_diff_vel -= 2 * np.pi
                elif angle_diff_vel < -np.pi:
                    angle_diff_vel += 2 * np.pi
                
                dt = (times[i] - times[i-1]) / 1e9
                curvature = abs(angle_diff_vel) / dt if dt > 0 else 0
                curvatures.append(curvature)
        
        curvature_smoothness = sum(curvature_changes)
        max_curvature = max(curvatures) if curvatures else 0.0
        avg_curvature = np.mean(curvatures) if curvatures else 0.0
        
        # Compute acceleration smoothness (velocity changes)
        acceleration_smoothness = 0.0
        if len(velocities) > 2:
            for i in range(1, len(velocities)):
                vx_prev = velocities[i-1][0]
                vy_prev = velocities[i-1][1]
                vx_curr = velocities[i][0]
                vy_curr = velocities[i][1]
                
                # Compute change in velocity
                dvx = vx_curr - vx_prev
                dvy = vy_curr - vy_prev
                acceleration_magnitude = np.sqrt(dvx*dvx + dvy*dvy)
                
                acceleration_smoothness += acceleration_magnitude
        
        return {
            'curvature_smoothness': curvature_smoothness,
            'acceleration_smoothness': acceleration_smoothness,
            'path_length': path_length,
            'avg_velocity': avg_velocity,
            'max_curvature': max_curvature,
            'avg_curvature': avg_curvature
        }
        
    except Exception as e:
        print(f"Warning: error when computing path smoothness: {e}")
        return {
            'curvature_smoothness': 0.0,
            'acceleration_smoothness': 0.0,
            'path_length': 0.0,
            'avg_velocity': 0.0,
            'max_curvature': 0.0,
            'avg_curvature': 0.0
        }

def load_metrics_data(data_dir):
    """Load metrics data for a single test."""
    metrics_data = {}
    
    # Base metrics file
    metrics_file = os.path.join(data_dir, "metrics.csv")
    if os.path.exists(metrics_file):
        try:
            df_metrics = pd.read_csv(metrics_file)
            metrics_data.update(df_metrics.iloc[0].to_dict())
        except Exception as e:
            print(f"Warning: failed to read {metrics_file}: {e}")
    
    # Subject metrics file (Tasks 1-5)
    subject_file = os.path.join(data_dir, "metrics_subject.csv")
    if os.path.exists(subject_file):
        try:
            df_subject = pd.read_csv(subject_file)
            # Add subject-related columns, avoiding duplicates
            for col in df_subject.columns:
                if col not in metrics_data:
                    metrics_data[col] = df_subject.iloc[0][col]
        except Exception as e:
            print(f"Warning: failed to read {subject_file}: {e}")
    
    # Zone metrics file
    zone_file = os.path.join(data_dir, "metrics_zone.csv")
    if os.path.exists(zone_file):
        try:
            df_zone = pd.read_csv(zone_file)
            # Add zone-related columns, avoiding duplicates
            for col in df_zone.columns:
                if col not in metrics_data:
                    metrics_data[col] = df_zone.iloc[0][col]
        except Exception as e:
            print(f"Warning: failed to read {zone_file}: {e}")
    
    # Compute path smoothness metrics
    odom_file = os.path.join(data_dir, "odom.csv")
    if os.path.exists(odom_file):
        try:
            smoothness_metrics = calculate_path_smoothness(odom_file)
            metrics_data.update(smoothness_metrics)
        except Exception as e:
            print(f"Warning: failed to compute path smoothness for {odom_file}: {e}")
    
    return metrics_data

def collect_all_data(base_dir):
    """Collect data for all tasks."""
    all_data = []  # Collect all data for Tasks 1-5
    
    # Iterate over all task directories
    for task_dir in sorted(glob.glob(os.path.join(base_dir, "task_*"))):
        task_name = os.path.basename(task_dir)
        print(f"Processing {task_name}...")
        
        # Iterate over all tests under this task
        for test_dir in sorted(glob.glob(os.path.join(task_dir, "task_*_test_*"))):
            test_name = os.path.basename(test_dir)
            jackal_dir = os.path.join(test_dir, "jackal")
            
            if not os.path.exists(jackal_dir):
                print(f"  Skipping {test_name}: missing jackal directory")
                continue
            
            print(f"  Processing {test_name}...")
            
            # Load data
            metrics_data = load_metrics_data(jackal_dir)
            
            if not metrics_data:
                print(f"  Skipping {test_name}: no valid data")
                continue
            
            # Add task and test identifiers
            metrics_data['task'] = task_name
            metrics_data['test'] = test_name
            
            # Add to all_data (Tasks 1-5)
            if task_name in ['task_1', 'task_2', 'task_3', 'task_4', 'task_5']:
                all_data.append(metrics_data)
            # Task 6 is skipped for now
    
    return all_data

def create_excel_file(all_data, output_file="tvss_nav_metrics.xlsx"):
    """Create the Excel file."""
    print(f"\nCreating Excel file: {output_file}")
    
    if not all_data:
        print("No data, skipping Excel file creation")
        return output_file
    
    df_all = pd.DataFrame(all_data)
    print(f"Tasks 1-5: {len(df_all)} records")
    
    # Create the Excel file
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_all.to_excel(writer, sheet_name='All_Tasks', index=False)
    
    # Apply color styling
    if OPENPYXL_AVAILABLE:
        apply_colors_to_excel(output_file, df_all)
    else:
        print("Skipping color styling (openpyxl not installed)")
    
    print(f"Excel file created: {output_file}")
    return output_file

def apply_colors_to_excel(file_path, df):
    """Apply color styling to the Excel file."""
    if not OPENPYXL_AVAILABLE:
        print("openpyxl not installed, cannot set color styling")
        return
        
    from openpyxl import load_workbook
    
    # Get color configuration
    task_colors = get_task_colors()
    task_font_colors = get_task_font_colors()
    
    # Load workbook
    wb = load_workbook(file_path)
    ws = wb['All_Tasks']
    
    # Find position of the task column
    task_col_idx = None
    for idx, col_name in enumerate(df.columns, 1):
        if col_name == 'task':
            task_col_idx = idx
            break
    
    if task_col_idx is None:
        print("Warning: could not find 'task' column, skipping color styling")
        return
    
    # Set colors for each row
    for row_idx in range(2, len(df) + 2):  # From row 2 (skip header row)
        task_value = ws.cell(row=row_idx, column=task_col_idx).value
        
        if task_value in task_colors:
            # Set background color
            fill_color = PatternFill(start_color=task_colors[task_value],
                                   end_color=task_colors[task_value],
                                   fill_type='solid')
            
            # Set font color
            font_color = Font(color=task_font_colors[task_value])
            
            # Apply to entire row
            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = fill_color
                cell.font = font_color
    
    # Save file
    wb.save(file_path)
    print("Color styling applied to Excel file")

def main():
    """Main function."""
    sub_dir_name = "tvss_nav"
    # sub_dir_name = "vlm_nav"
    # sub_dir_name = "vlm_social_nav"
    
    base_dir = f"data/{sub_dir_name}"
    print(f"Start collecting {sub_dir_name} data...")
    
    # Check data directory
    if not os.path.exists(base_dir):
        print(f"Error: could not find directory {base_dir}")
        return
    
    # Collect all data
    all_data = collect_all_data(base_dir)
    
    # Summary
    total_records = len(all_data)
    print("\nData collection finished:")
    print(f"  Tasks 1-5: {total_records} records")
    
    if total_records == 0:
        print("Warning: no valid data found")
        return
    
    # Create Excel file
    output_file = create_excel_file(all_data, output_file=f"{base_dir}/{sub_dir_name}_metrics_{format_datetime()}.xlsx")
    
    # Show column information
    print("\nColumn information:")
    if all_data:
        df = pd.DataFrame(all_data)
        print(f"  All_Tasks: {len(df.columns)} columns")
        print(f"    Column names: {list(df.columns)}")

if __name__ == "__main__":
    main()
